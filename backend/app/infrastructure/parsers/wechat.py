"""
WeChat Pay bill parser.

Supports:
- CSV exports
- XLSX exports from newer WeChat bill downloads
"""
import csv
import io
import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from app.domain.transaction.models import BillSource, RawTransaction, TransactionDirection
from app.infrastructure.parsers.base import BillParserAdapter
from app.infrastructure.parsers.registry import register

logger = logging.getLogger(__name__)

WECHAT_SIGNATURE = "微信支付账单明细"
WECHAT_HEADER = "交易时间"
WECHAT_HEADERS = {
    "交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)",
    "支付方式", "当前状态", "交易单号", "商户单号", "备注",
}

DIRECTION_MAP = {
    "支出": TransactionDirection.EXPENSE,
    "收入": TransactionDirection.INCOME,
    "/": TransactionDirection.TRANSFER,
    "不计收支": TransactionDirection.TRANSFER,
    "中性交易": TransactionDirection.TRANSFER,
}


class WeChatParser(BillParserAdapter):
    @property
    def source_type(self) -> str:
        return "wechat"

    def can_parse(self, content: str) -> bool:
        return WECHAT_SIGNATURE in content

    def can_parse_file(self, file_path: Path) -> bool:
        if file_path.suffix.lower() == ".xlsx":
            try:
                rows = self._read_xlsx_rows(file_path)
            except Exception:
                return False
            return any(self._is_header_row(row) for row in rows)
        return super().can_parse_file(file_path)

    def parse(self, content: str) -> list[RawTransaction]:
        content = content.lstrip("\ufeff")

        lines = content.splitlines()
        header_idx = None
        for i, line in enumerate(lines):
            if line.strip().startswith(WECHAT_HEADER):
                header_idx = i
                break

        if header_idx is None:
            logger.warning("WeChat parser: header row not found")
            return []

        csv_text = "\n".join(lines[header_idx:])
        reader = csv.DictReader(io.StringIO(csv_text))

        transactions: list[RawTransaction] = []
        for row in reader:
            try:
                tx = self._parse_row(row)
                if tx:
                    transactions.append(tx)
            except Exception as e:
                logger.debug("Skipping WeChat row: %s | error: %s", row, e)

        logger.info("WeChat parser: parsed %d transactions", len(transactions))
        return transactions

    def parse_file(self, file_path: Path) -> list[RawTransaction]:
        if file_path.suffix.lower() != ".xlsx":
            return super().parse_file(file_path)

        rows = self._read_xlsx_rows(file_path)
        header_idx = next((i for i, row in enumerate(rows) if self._is_header_row(row)), None)
        if header_idx is None:
            logger.warning("WeChat parser: XLSX header row not found")
            return []

        headers = [str(v).strip() for v in rows[header_idx]]
        transactions: list[RawTransaction] = []
        for values in rows[header_idx + 1:]:
            if not any(str(v).strip() for v in values):
                continue
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]}
            try:
                tx = self._parse_row(row)
                if tx:
                    transactions.append(tx)
            except Exception as e:
                logger.debug("Skipping WeChat XLSX row: %s | error: %s", row, e)

        logger.info("WeChat parser: parsed %d transactions from XLSX", len(transactions))
        return transactions

    def _is_header_row(self, row: list[object]) -> bool:
        normalized = {str(v).strip() for v in row if str(v).strip()}
        return WECHAT_HEADERS.issubset(normalized)

    def _read_xlsx_rows(self, file_path: Path) -> list[list[object]]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    def _parse_row(self, row: dict) -> RawTransaction | None:
        row = {str(k).strip(): ("" if v is None else str(v).strip()) for k, v in row.items() if k}

        direction_str = row.get("收/支", "/")
        direction = DIRECTION_MAP.get(direction_str, TransactionDirection.TRANSFER)

        status = row.get("当前状态", "")
        if "失败" in status or "退款" in status:
            return None

        amount_str = row.get("金额(元)", "0").replace("¥", "").replace(",", "").strip()
        amount = float(amount_str)

        dt_str = row.get("交易时间", "")
        transaction_at = self._parse_datetime(dt_str)

        merchant = row.get("交易对方", "")
        description = row.get("商品", "")

        return RawTransaction(
            source=BillSource.WECHAT,
            direction=direction,
            amount=amount,
            currency="CNY",
            merchant=merchant,
            description=description,
            transaction_at=transaction_at,
            raw_data=dict(row),
            external_id=row.get("交易单号") or None,
        )

    def _parse_datetime(self, value: str) -> datetime:
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass

        try:
            serial = float(value)
            return datetime(1899, 12, 30) + timedelta(days=serial)
        except ValueError:
            return datetime.now()


register(WeChatParser())
